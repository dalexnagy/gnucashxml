#!/usr/bin/python3
# -------------------------------------------------------------------------------------------
# Create report & spreadsheet of asset and liability accounts from GnuCash
#
# --- Change History ---
# Program version 0
# 2024-11-11 V1   - New (skeleton only)
# 2024-11-12 V1.1 - Added code to create Excel spreadsheet
# 2024-12-06 V1.5 - 1) Added function to format dollar amount
#                   2) Removed most of the 'skipped' messages
#                   3) Added code to skip future transactions w/message
#                   4) Changed spreadsheet title in line 1 & centered over 5 columns
# 2024-12-19 V1.6 - 1) Added code to get directory for workbook from command line args or use a default
#                   2) Moved book file name to a variable in the constants area
#                   3) Adjusted indent for account names
# 2024-12-20 V1.7 - Added code to set folder name if the command line parameter is empty
#
# System libraries
from datetime import date
from pathlib import Path
import argparse
# Openpyxl libraries
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Font, Alignment
# GnuCash library
import gnucashxml

# Set up and get any parameter
parser = argparse.ArgumentParser()
# Directory or Folder (optional)
parser.add_argument('directory', nargs='?', default=str(Path.home())+"/GnuCash/Reports/")
args = parser.parse_args()
#print("args.directory=", args.directory)

# CONSTANTS & Global Variables
Program_Version = "V1.6"
# Location of GnuCash file (book)
GnuCash_Book = "<your GnuCash Book File>.gnucash"
# Get current date and set other variables
today = date.today()
File_DatePrefix = "{:4d}-{:02d}-{:02d}".format(today.year, today.month, today.day)
# XLSX folder
if args.directory is None or args.directory == "":
    XLSX_Folder = str(Path.home())+"/GnuCash/Reports/"
else:
    XLSX_Folder = args.directory
#print("Report to be written to '{}'.".format(XLSX_Folder))

# Keep track of accounts already processed
processed_list = []
# Row number in spreadsheet
global sheet_row
# Total value
global total_value
total_value=0
# Accounts processed
global total_processed
total_processed = 0

# Format printed dollar amount
def formatDollarAmt(amt):
    if amt < 0:
        return "(${:,.2f})".format(abs(amt))
    else:
        return "${:,.2f}".format(amt)

# Process children recursively
def process_child(lvl, acc):
    global sheet_row
    global total_value
    global total_processed
    total_processed += 1
    child_type = ""
    processed_list.append(acc.name)
    lvl += 1
    for child in acc.children:
        if acc.actype != "EXPENSE" and acc.name+"-"+child.name not in processed_list:
            child_value = 0
            child_quantity = 0
            for split in child.splits:
                if split.transaction.date.date() <= today:
                    child_value += split.value
                    child_quantity += split.quantity
                    child_type = split.account.actype
                else:
                    print("--Skipping Future Trans in {} on {} ({}) for {}"
                          .format(child.name, split.transaction.date.date(), split.transaction.description, formatDollarAmt(split.value)))
            indent_space = "{:{}} ".format(" ", lvl)
            total_value += child_value
            if child_type in ["STOCK", "MUTUAL"] and child_quantity == 0:
                #print("--Skipping '{}' in '{}' account '{}' - 0 shares.".format(child.name, child_type, acc.name))
                pass
            elif child_value == 0 and len(child.children) == 0:
                #print("--Skipping '{}' in '{}' account '{}' - $0 value & 0 children.".format(child.name, child_type, acc.name))
                pass
            elif child_value != 0:
                #print("{:64} {:>12,.2f}".format(indent_space + child.name, child_value))
                ws["B{}".format(sheet_row)] = child.name
                ws["C{}".format(sheet_row)] = child_value
                ws["C{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                if child_type in ["STOCK", "MUTUAL"] and child_quantity != 0:
                    ws["D{}".format(sheet_row)] = child_quantity
                    ws["D{}".format(sheet_row)].number_format = '###,##0.0000'
                    ws["E{}".format(sheet_row)] = child_value / child_quantity
                    ws["E{}".format(sheet_row)].number_format = '"$"#,##0.0000'
                sheet_row += 1
            elif child_value == 0 and lvl > 0 and child_type not in ["STOCK", "MUTUAL"]:
                #print("{:64}".format(indent_space + child.name))
                ws["A{}".format(sheet_row)] = indent_space + child.name
                sheet_row += 1
            processed_list.append(child.name+"-"+acc.name)
            if len(acc.children) > 0:
                process_child(lvl, child)

# End of Functions

# Create spreadsheet and workbook
XLSX_filename = XLSX_Folder + "{}_AccountSummary.xlsx".format(File_DatePrefix)

wb = Workbook()
ws = wb.active
ws.title = "GnuCash Accounts"
DEFAULT_FONT.name = "FreeSans"
DEFAULT_FONT.size = 10

# Set column widths
ws.column_dimensions['A'].width = 12    # Date & Account Tree
ws.column_dimensions['B'].width = 48    # Account Name
ws.column_dimensions['C'].width = 20    # Value
ws.column_dimensions['D'].width = 16    # Quantity (Stocks & Mutual Funds)
ws.column_dimensions['E'].width = 16    # $/Share (Stocks & Mutual Funds)

sheet_row = 1
# Create Title Row
Todays_date = "{:02d}/{:02d}/{:4d}".format(today.month, today.day, today.year)
ws["A{}".format(sheet_row)] = "GnuCash Account Summary - Current to {}".format(Todays_date)
ws["A{}".format(sheet_row)].font = Font(bold=True)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
top_left_cell = ws['A1']
top_left_cell.alignment = Alignment(horizontal='center', vertical='center')

#ws["E{}".format(sheet_row)] = Todays_date#ws["E{}".format(sheet_row)].font = Font(bold=True)
#ws["E{}".format(sheet_row)].alignment = Alignment(horizontal='center')

sheet_row += 1
ws["B{}".format(sheet_row)] = "Account"
ws["B{}".format(sheet_row)].font = Font(bold=True)
ws["B{}".format(sheet_row)].alignment = Alignment(horizontal='center')

ws["C{}".format(sheet_row)] = "Balance"
ws["C{}".format(sheet_row)].font = Font(bold=True)
ws["C{}".format(sheet_row)].alignment = Alignment(horizontal='center')

ws["D{}".format(sheet_row)] = "Shares"
ws["D{}".format(sheet_row)].font = Font(bold=True)
ws["D{}".format(sheet_row)].alignment = Alignment(horizontal='center')

ws["E{}".format(sheet_row)] = "Share $"
ws["E{}".format(sheet_row)].font = Font(bold=True)
ws["E{}".format(sheet_row)].alignment = Alignment(horizontal='center')

sheet_row += 1

book = gnucashxml.from_filename(GnuCash_Book)

for account, children, splits in book.walk():

    if account.actype not in ["EXPENSE", "INCOME", "EQUITY"] and account.name != "Root Account":
        level = 0
        if account.name+"-"+account.parent.name not in processed_list and len(account.children) > 0:
            print("Processing '{}' from '{}' with {} children".format(account.name, account.parent.name, len(account.children)))
            process_child(level, account)
            processed_list.append(account.name + "-" + account.parent.name)
        #else:
        #    print("Skipping '{}' from '{}' - Processed earlier & 0 children".format(account.name, account.parent.name))

sheet_row += 1
ws["B{}".format(sheet_row)] = "TOTAL VALUE"
ws["B{}".format(sheet_row)].font = Font(bold=True)
ws["B{}".format(sheet_row)].alignment = Alignment(horizontal='center')

ws["C{}".format(sheet_row)] = total_value
ws["C{}".format(sheet_row)].font = Font(bold=True)
ws["C{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

print("\nCompleted processing {} accounts.".format(total_processed))

wb.save(XLSX_filename)
print("\nXLSX spreadsheet '{}' was created successfully.".format(XLSX_filename))
