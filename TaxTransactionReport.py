#!/usr/bin/python3
#----------------------------------------------------------------------------------------------
# Create report & spreadsheet of selected tax-related transactions from entire book of accounts
#
# --- Change History ---
# Program version 0
# 2024-12-24 V1   - New

Program_Version = "V1.0"

# System imports
import sys
from datetime import date
from pathlib import Path
# Openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Font, Alignment
# GnuCash Structure import
import gnucashxml

# CONSTANTS & Globals
# Get current date and set other variables
today = date.today()
File_DatePrefix = "{:4d}-{:02d}-{:02d}".format(today.year, today.month, today.day)
Generic_File_name = "{}_TaxRelatedTransactionReport".format(File_DatePrefix)
my_home = str(Path.home())
# Report folder
Report_Folder_name = my_home+"/GnuCash/Reports/"
# XLSX folder
Workbook_Folder_name = my_home+"/GnuCash/Reports/"
# Book File Location
book_file = "<your gnucash book file>.gnucash"
# END OF CONSTANTS

# Accounts to report
TaxRelatedAccounts_list = []
# Name of file containing list of accounts
account_list_file = "<your location>/TaxRelatedAccounts.txt"
# Open file and read each entry;  Add each to the list
with open(account_list_file) as f:
    for line in f:
        TaxRelatedAccounts_list.append(line.strip())
f.close()
#print("List of accounts:", TaxRelatedAccounts_list)

# Determine start and end dates for report
# if running in Jan to April, start date is 1/1 of prior year and end date is 12/31 of prior year
# else start date is 1/1 of current year and end date is today
if today.month <= 4:
    start_date = date(today.year - 1, 1, 1)
    end_date = date(today.year - 1, 12, 31)
else:
    start_date = date(today.year,1,1)
    end_date = today
report_title = "Tax-Related Transactions from {} through {} - {}".format(start_date.strftime("%m/%d/%Y"), end_date.strftime("%m/%d/%Y"), Program_Version)
print("Report of", report_title)
print("Accounts to inspect are in '{}'".format(account_list_file))
print("")

# Initialize transaction list
Found_Transactions_list = []

# -------------

# Format printed dollar amount
def formatDollarAmt(amt):
    if amt < 0:
        return "(${:,.2f})".format(abs(amt))
    else:
        return " ${:,.2f} ".format(amt)

# Format for value of 0 or None
def formatZeroNum(val):
    if val == 0:
        return "{:4}".format("")
    elif val is None:
        return ""
    else:
        return "{:4}".format(val)

# Format for value of None
def formatNone(val):
    if val is None:
        return ""
    else:
        return val

# Process split
def process_split(spl):
    if start_date <= spl.transaction.date.date() <= end_date:
        process_transaction(spl.account.name, spl.transaction)

# Process transaction and save in list
def process_transaction(acc_name, trans):
    save_account_name = acc_name
    save_value = 0
    for spl in trans.splits:
        if spl.account.name != acc_name and spl.value < 0:
            save_account_name = spl.account.name
        if spl.account.name == acc_name and spl.value > 0:
            save_value = spl.value
    Found_Transactions_list.append([acc_name, save_account_name, trans.date.strftime("%m/%d/%Y"), formatNone(trans.num), formatNone(trans.description),
                          save_value])

# Create a text report
def Create_Report(l):
    report_filename = Report_Folder_name + Generic_File_name + ".txt"

    # noinspection PyBroadException
    try:
        report_file = open(report_filename, 'w')
    except:
        print("Failed to open report file: {}".format(report_filename))
        print("Program fails with code 2")
        exit(2)

    report_file.write("{:^90}\n\n".format(report_title))
    # Column headings
    report_file.write("{:^28} {:^12} {:^6} {:^32} {:>12}\n"
                      .format("PAYMENT ACCOUNT", "DATE", "NUM", "DESCRIPTION", "AMOUNT"))
    report_file.write("{:^28} {:^12} {:^6} {:^32} {:>12}\n"
                      .format("---------------", "----", "---", "-----------", "------"))

    # Process the list of found transactions
    prior_tax_acct = ""
    total_value = 0

    for tax_acct, pay_acct, trx_date, trx_num, trx_desc, trx_value in l:
        if prior_tax_acct == "":
            prior_tax_acct = tax_acct
        if tax_acct != prior_tax_acct:
            report_file.write("{:28} {:^12} {:^6} {:>32} {:>12}\n\n"
                  .format(" ", " ", " ", "--- " + prior_tax_acct + " TOTAL:",  formatDollarAmt(total_value)))
            report_file.write("")
            total_value = 0
            prior_tax_acct = tax_acct
        if tax_acct == prior_tax_acct:
            report_file.write("{:28} {:^12} {:^6} {:<32} {:>12}\n"
                  .format(pay_acct, trx_date, formatNone(trx_num), formatNone(trx_desc), formatDollarAmt(trx_value)))
            total_value += trx_value
            prior_tax_acct = tax_acct

    report_file.write("{:28} {:^12} {:^6} {:>32} {:>12}\n\n"
          .format(" ", " ", " ", "--- " + prior_tax_acct + " TOTAL:",  formatDollarAmt(total_value)))
    report_file.close()
    print("Report Successfully Saved to '{}'".format(report_filename))

# Create spreadsheet of data
def Create_Workbook(l):  # Open workbook

    Workbook_filename =Workbook_Folder_name + Generic_File_name + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Related Transactions"
    DEFAULT_FONT.name = "FreeSans"
    DEFAULT_FONT.size = 10

    # Set column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 16

    # Create workbook title
    ws["A1"] = report_title
    ws["A1"].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Create Heading Row
    ws['A3'] = "TAX ACCOUNT"
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['B3'] = "PAY ACCOUNT"
    ws['B3'].font = Font(bold=True)
    ws['B3'].alignment = Alignment(horizontal='center')

    ws['C3'] = "DATE"
    ws['C3'].font = Font(bold=True)
    ws['C3'].alignment = Alignment(horizontal='center')

    ws['D3'] = "NUM"
    ws['D3'].font = Font(bold=True)
    ws['D3'].alignment = Alignment(horizontal='center')

    ws['E3'] = "DESCRIPTION"
    ws['E3'].font = Font(bold=True)
    ws['E3'].alignment = Alignment(horizontal='center')

    ws['F3'] = "AMOUNT"
    ws['F3'].font = Font(bold=True)
    ws['F3'].alignment = Alignment(horizontal='center')

    sum_start_row = 4
    sheet_row = 4

    prior_tax_acct = ""
    total_value = 0

    for tax_acct, pay_acct, trx_date, trx_num, trx_desc, trx_value in l:
        if prior_tax_acct == "":
            prior_tax_acct = tax_acct

        if tax_acct != prior_tax_acct:
            ws["E{}".format(sheet_row)] = "TOTAL:"
            ws["E{}".format(sheet_row)].alignment = Alignment(horizontal='right')
            ws["E{}".format(sheet_row)].font = Font(bold=True)
            ws["F{}".format(sheet_row)] = "=SUM(F{}:F{})".format(sum_start_row, sheet_row - 1)
            ws["F{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            ws["F{}".format(sheet_row)].font = Font(bold=True)
            sum_start_row = sheet_row + 2
            sheet_row += 2
            total_value = 0
            prior_tax_acct = tax_acct

        if tax_acct == prior_tax_acct:
            ws["A{}".format(sheet_row)] = tax_acct
            ws["A{}".format(sheet_row)].alignment = Alignment(horizontal='left')
            ws["B{}".format(sheet_row)] = pay_acct
            ws["B{}".format(sheet_row)].alignment = Alignment(horizontal='left')
            ws["C{}".format(sheet_row)] = trx_date
            ws["C{}".format(sheet_row)].alignment = Alignment(horizontal='center')
            ws["D{}".format(sheet_row)] = formatNone(trx_num)
            ws["D{}".format(sheet_row)].alignment = Alignment(horizontal='center')
            ws["E{}".format(sheet_row)] = trx_desc
            ws["E{}".format(sheet_row)].alignment = Alignment(horizontal='left')
            ws["F{}".format(sheet_row)] = trx_value
            ws["F{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            sheet_row += 1
            total_value += trx_value
            prior_tax_acct = tax_acct

    ws["E{}".format(sheet_row)] = "TOTAL:"
    ws["E{}".format(sheet_row)].alignment = Alignment(horizontal='right')
    ws["E{}".format(sheet_row)].font = Font(bold=True)
    ws["F{}".format(sheet_row)] = "=SUM(F{}:F{})".format(sum_start_row, sheet_row - 1)
    ws["F{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    ws["F{}".format(sheet_row)].font = Font(bold=True)

    try:
        wb.save(Workbook_filename)
    except:
        print("Could not save XLSX file: {}".format(sys.exc_info()[0]))
        exit(3)

    print("Workbook successfully saved to '{}'".format(Workbook_filename))

#-------------

book = gnucashxml.from_filename(book_file")

for account, children, splits in book.walk():
    if len(splits) > 0 and account.name in TaxRelatedAccounts_list:
        #print("ROOT.{:48} PARENT: {:32} TYPE: {:16} #CHILDREN:{:3} #SPLITS:{:4}".format(account.name, formatNone(account.parent.name), formatNone(account.actype), len(children), len(splits)))
        for split in account.splits:
            process_split(split)

# After finding transactions, sort, then summarize

# Sort by tax-account name then by date
Found_Transactions_list.sort(key=lambda x: (x[0], x[2]))

# Create text report
Create_Report(Found_Transactions_list)

# Create workbook
Create_Workbook(Found_Transactions_list)

