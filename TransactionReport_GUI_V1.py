#!/usr/bin/python3

#-------------------------------------------------------------------------------------------
# Create report & spreadsheet of selected transactions from a selected account
#
# --- Change History ---
# Program version 0
# 2024-12-11 V1   - New
# 2024-12-12 V1.1 - Added checkbox to avoid calculating & using a prior balance
# 2024-12-13 V1.2 - Minor text changes on report title & headings;
#                   Added part of selected account name to the file names;
#                   Fixed issue with an empty description field

Program_Version = "V1.2"

# System imports
import sys
import os
import re
from datetime import date
from pathlib import Path
# Openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Font, Alignment
# GUI imports
from PyQt5 import QtCore, QtWidgets, uic
from PyQt5.QtWidgets import (QApplication, QTableWidgetItem, QHeaderView)
# GnuCash Structure import
import gnucashxml

# QT Class - AlignDelegates
class CenterAlignDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super(CenterAlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignCenter
class RightAlignDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super(RightAlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignRight

# set path from my own imports
sys.path.insert(0, '/home/dave/Python/_Configs')

# CONSTANTS & Globals
# Globals
# Get current date and set other variables
today = date.today()
File_DatePrefix = "{:4d}-{:02d}-{:02d}".format(today.year, today.month, today.day)
#Generic_File_name = "{}_TransactionReport".format(File_DatePrefix)
my_home = str(Path.home())
# Report folder
Report_Folder_name = my_home+"/GnuCash/Reports/"
# XLSX folder
Workbook_Folder_name = my_home+"/GnuCash/Reports/"
# Command to print the report
lpr_cmd="lpr -o orientation-requested=4 -o cpi=12 -o page-left=36 -o page-top=36"
# Book File Location
book_file = "/home/dave/GnuCash/NagyFamily2024.gnucash"
# Initialize transaction list
transaction_list = []
# List of master accounts to keep in list:
candidate_types = ["BANK", "CREDIT", "EXPENSE", "INCOME"]
# END OF CONSTANTS

# Global Functions
# Format printed dollar amount
def formatDollarAmt(amt):
    if amt < 0:
        return "(${:,.2f})".format(abs(amt))
    else:
        return " ${:,.2f} ".format(amt)

# Format if no value
def formatZeroNone(val):
    if val == 0 or val is None:
        return ""
    else:
        return val

# Process account
def process_account(acc):
    if len(acc.splits) > 0 or acc.actype == "EXPENSE":
        account_list.append(acc.name)
    if len(acc.children) > 0:
        for ch in acc.children:
            #print("process_account: Child: ", ch.name, " #Children:", len(ch.children), " # Splits:", len(ch.splits))
            process_account(ch)

# Remove blanks and special characters from a string
def clean_string(s):
    s = re.sub(r'\W+', '', s)  # remove non-word characters (blanks, punctuation, etc.)
    return s

# End of global functions

# Global variables
# Initialize balance variables
prior_balance = 0
prior_balance_date = date(2000,1,1)
# Initialize transaction list
#transaction_list = []
# End of Global variables

print("initialize: Open Book & Get all Accounts")
book = gnucashxml.from_filename(book_file)
account_list = []
# Build account list
for account, children, splits in book.walk():
    #print("initialize: Account:",account.name, " # Children: ", len(account.children), " # Splits:", len(account.splits))
    if account.actype in candidate_types:
        if len(account.children) > 0:
            for child in account.children:
                #print("initialize: Child: ", child.name, " # Splits:", len(child.splits))
                process_account(child)
        elif len(account.splits) > 0:
            #print("initialize: Account:", account.name, " # Children: ", len(account.children), " # Splits:", len(account.splits))
            process_account(account)

account_list = sorted(set(account_list))

print("initialize: Book is open and account list complete and sorted")
#print("initialize: account_list:", account_list)

# Open front page window
Ui_MainWindow, QtBaseClass = uic.loadUiType("TransactionReport.ui")

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    def __init__(self):
        # QT Initialization of main window
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)

        # Set title
        self.Title_label.setText("GnuCash Transaction Report " + Program_Version)

        # Set actions & initial status for buttons
        self.Process_button.clicked.connect(self.Process_Request)
        self.Report_Save_button.clicked.connect(self.Create_Report)
        self.Report_Save_button.setEnabled(False)
        self.Print_Report_button.clicked.connect(self.Print_Report)
        self.Print_Report_button.setEnabled(False)
        self.Workbook_Save_button.clicked.connect(self.Create_Workbook)
        self.Workbook_Save_button.setEnabled(False)
        self.Exit_button.clicked.connect(self.exitNow)
        
        # Load Account selection
        self.Account_box.addItems(account_list)

        # Set screen fields to initial values
        self.Book_File_label.setText(book_file)
        self.End_dateEdit.setDate(today)
        #self.Report_File_entry.setText(Report_Folder_name + Generic_File_name+".txt")
        #self.Workbook_File_entry.setText(Workbook_Folder_name + Generic_File_name + ".xlsx")

    def PopulateTable(self, trx_list):

        # Set titles
        self.Transaction_List.setHorizontalHeaderLabels(["Date", "Num", "Description", "Amount"])

        # Align data in selected columns
        CtrDelegate = CenterAlignDelegate(self.Transaction_List)
        RtDelegate = RightAlignDelegate(self.Transaction_List)
        self.Transaction_List.setItemDelegateForColumn(0, CtrDelegate)
        self.Transaction_List.setItemDelegateForColumn(1, CtrDelegate)
        self.Transaction_List.setItemDelegateForColumn(3, RtDelegate)
        # Set table to auto adjust
        self.Transaction_List.setSizeAdjustPolicy(
            QtWidgets.QAbstractScrollArea.AdjustToContents)
        # Set column widths
        header = self.Transaction_List.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        # Load table with data from database tables
        table_row = 0

        for trx_date, trx_num, trx_desc, trx_amt in trx_list:
            self.Transaction_List.resizeColumnsToContents()
            self.Transaction_List.insertRow(table_row)
            self.Transaction_List.setItem(table_row, 0, QTableWidgetItem(trx_date.strftime("%m/%d/%Y")))
            self.Transaction_List.setItem(table_row, 1, QTableWidgetItem(formatZeroNone(trx_num)))
            self.Transaction_List.setItem(table_row, 2, QTableWidgetItem(trx_desc))
            self.Transaction_List.setItem(table_row, 3, QTableWidgetItem(formatDollarAmt(trx_amt)))
            #print("PopulateTable: table_row=", table_row, trx_date, trx_num, trx_desc, trx_amt)
            table_row += 1

        # Enable report and workbook buttons
        self.Report_Save_button.setEnabled(True)
        self.Workbook_Save_button.setEnabled(True)

    def Process_Request(self):
        global prior_balance
        global prior_balance_date
        global transaction_list

        # Clear Table, list and other values
        self.Transaction_List.clear()
        self.Transaction_List.setRowCount(0)
        self.Transaction_List.clearContents()
        self.Transaction_List.clearSpans()

        transaction_list.clear()
        prior_balance = 0
        prior_balance_date = date(2000,1,1)

        sel_account = self.Account_box.currentText()
        sel_start = self.Start_dateEdit.date().toPyDate()
        sel_end = self.End_dateEdit.date().toPyDate()

        for acc2, child2, split2 in book.walk():
            if acc2.name == sel_account:
                #print("Processing:", account.name)
                for split3 in acc2.splits:
                    self.process_split(split3, sel_start, sel_end)

        if not self.No_Prior_Balance_checkbox.isChecked():
            # Add prior balance to list
            transaction_list.append([prior_balance_date, 0, "{:>80}".format("-- PRIOR BALANCE --"), prior_balance])
        # sort list by date
        transaction_list.sort(key=lambda x: (x[0], x[3]))  # Sort by date then value
        # Calculate current balance and put in transaction list
        current_balance = 0
        for trx_date, trx_num, trx_desc, trx_value in transaction_list:
            current_balance += trx_value
        transaction_list.append([prior_balance_date, 0, "{:>80}".format("-- CURRENT BALANCE/TOTAL --"), current_balance])

        account_str = clean_string(sel_account)
        account_File_name = "{}_{}_Transactions".format(File_DatePrefix, account_str[:16])
        self.Report_File_entry.setText(Report_Folder_name + account_File_name + ".txt")
        self.Workbook_File_entry.setText(Workbook_Folder_name + account_File_name + ".xlsx")

        self.PopulateTable(transaction_list)

    # noinspection PyMethodMayBeStatic
    def process_split(self, split, sta_date, end_date):
        global prior_balance
        global prior_balance_date
        global transaction_list
        if split.transaction.date.date() < sta_date:
            prior_balance += split.value
            #print("process_split: Prior Balance:", prior_balance)
            if split.transaction.date.date() > prior_balance_date:
                prior_balance_date = split.transaction.date.date()
        elif split.transaction.date.date() >= sta_date and split.transaction.date.date() <= end_date:
            transaction_list.append(
                [split.transaction.date.date(), split.transaction.num, split.transaction.description, split.value])
            #print("process_split: Add ", split.transaction.date.date(), split.transaction.num, split.transaction.description, split.value)

    def Create_Report(self):
        self.Report_Msg_label.clear()
        self.Report_Msg_label.setStyleSheet("background-color: white; color: blue;")

        report_filename = self.Report_File_entry.text()

        # noinspection PyBroadException
        try:
            report_file = open(report_filename,'w')
        except:
            self.ReportMsg_label.setStyleSheet("background-color: yellow; color: red;")
            self.ReportMsg_label.setText("Could not open file: {}".format(sys.exc_info()[0]))
            self.SaveReport_checkbox.setChecked(False)
            self.PrintReport_checkbox.setChecked(False)
            return

        title = "Transaction Report from Account'" + self.Account_box.currentText() + "'"
        report_file.write("{:^112}\n\n".format(title))
        # Column headings
        report_file.write("{:^12} {:^6} {:80} {:^14}\n"
                          .format("DATE", "NUM", "DESCRIPTION", "AMOUNT"))
        report_file.write("{:^12} {:^6} {:80} {:^14}\n"
                          .format("----", "---", "-----------", "------"))

        for trx_date, trx_num, trx_desc, trx_amt in transaction_list:
            #print("Create_Report: trx_date=", trx_date, " trx_num=", trx_num, "From FormatZeroNone:'", formatZeroNone(trx_num), "'", trx_desc, trx_amt)
            report_file.write("{:12} {:^6} {:80} {:>14}\n"
                              .format(trx_date.strftime("%m/%d/%Y"), formatZeroNone(trx_num), formatZeroNone(trx_desc), formatDollarAmt(trx_amt)))

        report_file.close()
        self.Report_Msg_label.setText("Successfully Saved!")

        # Enable print report button
        self.Print_Report_button.setEnabled(True)
        
    def Print_Report(self):
        report_filename = self.Report_File_entry.text()
        os.system(lpr_cmd+" {}".format(report_filename))

    def Create_Workbook(self):# Open workbook

        self.Workbook_Msg_label.clear()
        self.Workbook_Msg_label.setStyleSheet("background-color: white; color: blue;")

        Workbook_filename = self.Workbook_File_entry.text()

        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Report"
        DEFAULT_FONT.name = "FreeSans"
        DEFAULT_FONT.size = 10

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 16

        # Create workbook title
        ws["A1"] = "Transaction Report from '" + self.Account_box.currentText() + "'"
        ws["A1"].font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        # Create Heading Row
        ws['A3'] = "DATE"
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws['B3'] = "NUM"
        ws['B3'].font = Font(bold=True)
        ws['B3'].alignment = Alignment(horizontal='center')

        ws['C3'] = "DESCRIPTION"
        ws['C3'].font = Font(bold=True)
        ws['C3'].alignment = Alignment(horizontal='center')

        ws['D3'] = "AMOUNT"
        ws['D3'].font = Font(bold=True)
        ws['D3'].alignment = Alignment(horizontal='center')

        sheet_row = 4

        #transactionList, count = self.fetchRows(int(self.TransactionCount.text()))

        for trx_date, trx_num, trx_desc, trx_amt in transaction_list:
            ws["A{}".format(sheet_row)] = trx_date
            ws["A{}".format(sheet_row)].alignment = Alignment(horizontal='center')
            ws["B{}".format(sheet_row)] = formatZeroNone(trx_num)
            ws["B{}".format(sheet_row)].alignment = Alignment(horizontal='center')
            ws["C{}".format(sheet_row)] = trx_desc
            ws["D{}".format(sheet_row)] = trx_amt
            ws["D{}".format(sheet_row)].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            sheet_row += 1

        # noinspection PyBroadException
        try:
            wb.save(Workbook_filename)
        except:
            self.Workbook_Msg_label.setStyleSheet("background-color: yellow; color: red;")
            self.Workbook_Msg_label.setText("Could not save XLSX file: {}".format(sys.exc_info()[0]))
            self.DoSpreadsheet_checkbox.setChecked(False)
            return

        self.Workbook_Msg_label.setText("Successfully Saved!")

    # noinspection PyMethodMayBeStatic
    def exitNow(self):
        sys.exit(0)

# Start the main window
def main():
    app = QApplication(sys.argv)
    # Instantiate main window
    w = MainWindow()
    w.show()

    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
